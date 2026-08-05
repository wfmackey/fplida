#!/usr/bin/env python3
"""Build packaged LFS metadata from ABS LLFS and COE data item lists.

The source workbook is the public "66020 LLFS Data Item List" workbook. It
stores one data item per row where the Survey column is ``LFS`` and category
codes in the following rows. This script keeps the workbook variable list as
the source of truth, expands compact identifier ranges such as
``REPWTG01 - REPWTG30``, and adds the synthetic fplida linkage key.

The Characteristics of Employment (COE/COES) file is a supplementary LLFS
DataLab file with a separate data item list. It has embedded flag rows and
``Values`` blocks inside a single conceptual item block, so it needs a parser
that treats flag identifiers as variables when they differ from the parent
identifier and treats ``Values`` rows as code-domain section markers.
"""

from __future__ import annotations

import argparse
import csv
import re
import tempfile
import urllib.request
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


SOURCE_WORKBOOK = "66020 LLFS Data Item List (updated Apr 2026).xlsx"
COES_SOURCE_URL = (
    "https://www.abs.gov.au/statistics/microdata-tablebuilder/"
    "available-microdata-tablebuilder/characteristics-employment-australia/"
    "COE%202025%20Data%20Item%20List%20DataLab.xlsx"
)
EXCLUDED_SHEETS = {"Contents"}
HEADER_ROW = 8


def clean(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def collection_cadence(frequency: str) -> str:
    lower = frequency.casefold()
    if lower.startswith("month"):
        return "monthly"
    if lower.startswith("quarter"):
        return "quarterly"
    if lower.startswith("annual"):
        return "annual"
    if lower.startswith("biennial"):
        return "biennial"
    return "static"


def expand_identifier(identifier: str, following_identifier: str = "") -> list[str]:
    joined = f"{identifier}{following_identifier}".replace(" ", "")
    match = re.fullmatch(r"([A-Z]+)(\d+)-([A-Z]+)?(\d+)", joined)
    if not match:
        return [identifier.replace(" ", "")]

    prefix1, start_text, prefix2, end_text = match.groups()
    prefix2 = prefix2 or prefix1
    if prefix1 != prefix2:
        return [identifier.replace(" ", "")]

    start = int(start_text)
    end = int(end_text)
    width = max(len(start_text), len(end_text))
    if start > end:
        return [identifier.replace(" ", "")]
    return [f"{prefix1}{idx:0{width}d}" for idx in range(start, end + 1)]


def header_map(ws) -> dict[str, int]:
    header = next(ws.iter_rows(min_row=HEADER_ROW, max_row=HEADER_ROW))
    return {
        clean(cell.value): idx
        for idx, cell in enumerate(header)
        if clean(cell.value)
    }


def rows(ws) -> Iterable[tuple[int, list[str]]]:
    for row_number, row in enumerate(
        ws.iter_rows(min_row=HEADER_ROW + 1), start=HEADER_ROW + 1
    ):
        yield row_number, [clean(cell.value) for cell in row]


def cell(cells: list[str], idx: int) -> str:
    if idx < 0 or idx >= len(cells):
        return ""
    return cells[idx]


def is_code(value: str) -> bool:
    return value == "..." or re.fullmatch(r"-?\d+(?:\.\d+)?", value) is not None


def read_variables(workbook: Path) -> tuple[list[dict[str, str]], list[dict[str, str]]]:
    wb = load_workbook(workbook, read_only=True, data_only=True)
    variables: list[dict[str, str]] = [
        {
            "identifier": "SYNTHETIC_AEUID",
            "label": "Synthetic MADIP Person ID",
            "sheet": "fplida linkage",
            "source_row": "",
            "level": "Person",
            "survey": "LFS",
            "frequency": "Not applicable",
            "population": "All",
            "collection_cadence": "static",
            "source_workbook": workbook.name,
        }
    ]
    values: list[dict[str, str]] = []

    for sheet_name in wb.sheetnames:
        if sheet_name in EXCLUDED_SHEETS:
            continue
        ws = wb[sheet_name]
        headers = header_map(ws)
        identifier_col = headers["Identifier"]
        level_col = headers["Level"]
        survey_col = headers["Survey"]
        frequency_col = headers["Frequency"]
        population_col = headers.get("Population", -1)
        materialised = list(rows(ws))
        for idx, (row_number, cells) in enumerate(materialised):
            label = cell(cells, 2)
            identifier = cell(cells, identifier_col)
            level = cell(cells, level_col)
            survey = cell(cells, survey_col)
            frequency = cell(cells, frequency_col)
            population = cell(cells, population_col)
            if survey != "LFS" or not identifier:
                continue

            following_identifier = ""
            if identifier.endswith("-") and idx + 1 < len(materialised):
                following_identifier = cell(materialised[idx + 1][1], identifier_col)

            identifiers = expand_identifier(identifier, following_identifier)
            for expanded in identifiers:
                variables.append(
                    {
                        "identifier": expanded,
                        "label": label,
                        "sheet": sheet_name,
                        "source_row": str(row_number),
                        "level": level,
                        "survey": survey,
                        "frequency": frequency,
                        "population": population,
                        "collection_cadence": collection_cadence(frequency),
                        "source_workbook": workbook.name,
                    }
                )

            value_order = 0
            for value_row_number, value_cells in materialised[idx + 1 :]:
                next_identifier = cell(value_cells, identifier_col)
                next_survey = cell(value_cells, survey_col)
                if next_survey == "LFS" and next_identifier:
                    break
                code = cell(value_cells, 2)
                value_label = cell(value_cells, 3)
                if not code and not value_label:
                    continue
                if not code or not is_code(code):
                    continue
                if value_label.lower().startswith(("note:", "excludes", "includes")):
                    continue

                value_order += 1
                for expanded in identifiers:
                    values.append(
                        {
                            "identifier": expanded,
                            "code": code,
                            "value_label": value_label,
                            "frequency_note": cell(value_cells, frequency_col),
                            "sheet": sheet_name,
                            "source_row": str(value_row_number),
                            "value_order": str(value_order),
                            "is_placeholder": str(code == "..." and value_label == "...").upper(),
                            "source_workbook": workbook.name,
                        }
                    )

    return variables, values


def is_top_level_data_item(cells: list[str], headers: dict[str, int]) -> bool:
    identifier = cell(cells, headers["Identifier"])
    label = cell(cells, 2)
    survey = cell(cells, headers["Survey"])
    return bool(identifier and label and survey in {"LFS", "COE"})


def read_coes_variables(workbook: Path) -> tuple[list[dict[str, str]], list[dict[str, str]]]:
    wb = load_workbook(workbook, read_only=True, data_only=True)
    variables: list[dict[str, str]] = [
        {
            "identifier": "SYNTHETIC_AEUID",
            "label": "Synthetic MADIP Person ID",
            "sheet": "fplida linkage",
            "source_row": "",
            "level": "Person",
            "survey": "fplida",
            "frequency": "Not applicable",
            "platform": "DataLab",
            "collection_cadence": "static",
            "source_workbook": workbook.name,
        }
    ]
    values: list[dict[str, str]] = []

    def add_values(
        identifiers: list[str],
        sheet_name: str,
        materialised: list[tuple[int, list[str]]],
        start: int,
        stop: int,
        frequency_col: int,
    ) -> None:
        value_order = 0
        for value_row_number, value_cells in materialised[start:stop]:
            code = cell(value_cells, 2)
            value_label = cell(value_cells, 3)
            if not code and not value_label:
                continue
            if not code or not is_code(code):
                continue
            if value_label.lower().startswith(("note:", "excludes", "includes")):
                continue

            value_order += 1
            for expanded in identifiers:
                values.append(
                    {
                        "identifier": expanded,
                        "code": code,
                        "value_label": value_label,
                        "frequency_note": cell(value_cells, frequency_col),
                        "sheet": sheet_name,
                        "source_row": str(value_row_number),
                        "value_order": str(value_order),
                        "is_placeholder": str(
                            code == "..." and value_label == "..."
                        ).upper(),
                        "source_workbook": workbook.name,
                    }
                )

    for sheet_name in wb.sheetnames:
        if sheet_name in EXCLUDED_SHEETS:
            continue
        ws = wb[sheet_name]
        headers = header_map(ws)
        identifier_col = headers["Identifier"]
        level_col = headers["Level"]
        survey_col = headers["Survey"]
        frequency_col = headers["Frequency"]
        platform_col = headers.get("Platform", headers.get("Population", -1))
        materialised = list(rows(ws))

        idx = 0
        while idx < len(materialised):
            row_number, cells = materialised[idx]
            if not is_top_level_data_item(cells, headers):
                idx += 1
                continue

            label = cell(cells, 2)
            identifier = cell(cells, identifier_col)
            level = cell(cells, level_col)
            survey = cell(cells, survey_col)
            frequency = cell(cells, frequency_col)
            platform = cell(cells, platform_col)
            following_identifier = ""
            if identifier.endswith("-") and idx + 1 < len(materialised):
                following_identifier = cell(materialised[idx + 1][1], identifier_col)
            identifiers = expand_identifier(identifier, following_identifier)

            for expanded in identifiers:
                variables.append(
                    {
                        "identifier": expanded,
                        "label": label,
                        "sheet": sheet_name,
                        "source_row": str(row_number),
                        "level": level,
                        "survey": survey,
                        "frequency": frequency,
                        "platform": platform,
                        "collection_cadence": collection_cadence(frequency),
                        "source_workbook": workbook.name,
                    }
                )

            next_idx = idx + 1
            while next_idx < len(materialised) and not is_top_level_data_item(
                materialised[next_idx][1], headers
            ):
                next_idx += 1

            markers = [
                marker_idx
                for marker_idx in range(idx + 1, next_idx)
                if cell(materialised[marker_idx][1], 2) in {"Flag", "Values"}
                and cell(materialised[marker_idx][1], identifier_col)
            ]
            if not markers:
                add_values(
                    identifiers,
                    sheet_name,
                    materialised,
                    idx + 1,
                    next_idx,
                    frequency_col,
                )
                idx = next_idx
                continue

            marker_pos = idx + 1
            while marker_pos < next_idx:
                marker_cells = materialised[marker_pos][1]
                marker_label = cell(marker_cells, 2)
                marker_identifier = cell(marker_cells, identifier_col)
                if marker_label == "Flag" and marker_identifier:
                    marker_identifiers = expand_identifier(marker_identifier)
                    if not same_identifier_set(marker_identifiers, identifiers):
                        marker_frequency = cell(marker_cells, frequency_col) or frequency
                        for expanded in marker_identifiers:
                            variables.append(
                                {
                                    "identifier": expanded,
                                    "label": f"{label} - flag",
                                    "sheet": sheet_name,
                                    "source_row": str(materialised[marker_pos][0]),
                                    "level": level,
                                    "survey": survey,
                                    "frequency": marker_frequency,
                                    "platform": platform,
                                    "collection_cadence": collection_cadence(
                                        marker_frequency
                                    ),
                                    "source_workbook": workbook.name,
                                }
                            )
                    stop = marker_pos + 1
                    while stop < next_idx:
                        stop_cells = materialised[stop][1]
                        if (
                            cell(stop_cells, 2) in {"Flag", "Values"}
                            and cell(stop_cells, identifier_col)
                        ):
                            break
                        stop += 1
                    add_values(
                        marker_identifiers,
                        sheet_name,
                        materialised,
                        marker_pos + 1,
                        stop,
                        frequency_col,
                    )
                    marker_pos = stop
                    continue
                if marker_label == "Values" and marker_identifier:
                    marker_identifiers = expand_identifier(marker_identifier)
                    stop = marker_pos + 1
                    while stop < next_idx:
                        stop_cells = materialised[stop][1]
                        if (
                            cell(stop_cells, 2) in {"Flag", "Values"}
                            and cell(stop_cells, identifier_col)
                        ):
                            break
                        stop += 1
                    add_values(
                        marker_identifiers,
                        sheet_name,
                        materialised,
                        marker_pos + 1,
                        stop,
                        frequency_col,
                    )
                    marker_pos = stop
                    continue
                marker_pos += 1

            idx = next_idx

    return variables, values


def same_identifier_set(left: list[str], right: list[str]) -> bool:
    return sorted(left) == sorted(right)


def write_csv(path: Path, rows_: list[dict[str, str]], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows_)


def resolve_input(path_or_url: str, default_filename: str) -> Path:
    if re.match(r"https?://", path_or_url):
        out_path = Path(tempfile.gettempdir()) / default_filename
        urllib.request.urlretrieve(path_or_url, out_path)
        return out_path
    return Path(path_or_url).expanduser()


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--input",
        default=str(Path.home() / "Downloads" / SOURCE_WORKBOOK),
        help="Path to the ABS LLFS data item list workbook.",
    )
    parser.add_argument(
        "--output-dir",
        default="inst/extdata/llfs",
        help="Directory for packaged LLFS metadata CSVs.",
    )
    parser.add_argument(
        "--coes-input",
        default=COES_SOURCE_URL,
        help="Path or URL to the COE/COES DataLab data item list workbook.",
    )
    args = parser.parse_args()

    workbook = Path(args.input).expanduser()
    if not workbook.exists():
        raise FileNotFoundError(workbook)
    coes_workbook = resolve_input(
        args.coes_input,
        "COE_2025_Data_Item_List_DataLab.xlsx",
    )
    if not coes_workbook.exists():
        raise FileNotFoundError(coes_workbook)

    output_dir = Path(args.output_dir)
    variables, values = read_variables(workbook)
    coes_variables, coes_values = read_coes_variables(coes_workbook)
    write_csv(
        output_dir / "llfs_variables.csv",
        variables,
        [
            "identifier",
            "label",
            "sheet",
            "source_row",
            "level",
            "survey",
            "frequency",
            "population",
            "collection_cadence",
            "source_workbook",
        ],
    )
    write_csv(
        output_dir / "llfs_values.csv",
        values,
        [
            "identifier",
            "code",
            "value_label",
            "frequency_note",
            "sheet",
            "source_row",
            "value_order",
            "is_placeholder",
            "source_workbook",
        ],
    )
    write_csv(
        output_dir / "coes_variables.csv",
        coes_variables,
        [
            "identifier",
            "label",
            "sheet",
            "source_row",
            "level",
            "survey",
            "frequency",
            "platform",
            "collection_cadence",
            "source_workbook",
        ],
    )
    write_csv(
        output_dir / "coes_values.csv",
        coes_values,
        [
            "identifier",
            "code",
            "value_label",
            "frequency_note",
            "sheet",
            "source_row",
            "value_order",
            "is_placeholder",
            "source_workbook",
        ],
    )

    print(
        "Wrote LLFS metadata: "
        f"{len(variables)} variables and {len(values)} coded value rows."
    )
    print(
        "Wrote COES metadata: "
        f"{len(coes_variables)} variables and {len(coes_values)} coded value rows."
    )


if __name__ == "__main__":
    main()
